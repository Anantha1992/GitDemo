import openpyxl

book = openpyxl.load_workbook("C:\\Users\\ANANTHA\\Desktop\\PythonDemo.xlsx")

sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
#insert empty dic
Dict={}
# to write data

sheet.cell(row=2, column=2).value = "Ananth"

print(sheet.cell(row=2, column=2).value)

# to print no of rows and columns

print(sheet.max_row)
print(sheet.max_column)

# shortct to print values
print(sheet['A2'].value)

# to iterate and print all avalue in rows and columns
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)
# to print the test case data we needed

for i in range(1,sheet.max_row+1):
    if (sheet.cell(row=i, column=1).value)=="Testcase2":
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)

#to get data from excel and store it as dic in console
for i in range(1,sheet.max_row+1):
    if (sheet.cell(row=i, column=1).value)=="Testcase2":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i, column=j).value
print(Dict)
